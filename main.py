import os
from telegram import (
    InlineKeyboardButton, InlineKeyboardMarkup,
    KeyboardButton, ReplyKeyboardMarkup, Update
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters
)
from openpyxl import Workbook, load_workbook

# 🔧 تابع ذخیره در فایل Excel
def save_to_excel(name, phone=None, location=None):
    filename = "data.xlsx"

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["نام", "شماره تماس", "لوکیشن"])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    location_str = f"{location.latitude}, {location.longitude}" if location else ""
    ws.append([name, phone or "", location_str])
    wb.save(filename)

# 🎉 پیام خوش‌آمدگویی
welcome_text = (
    "👋 به ربات *مهدی حسنلو* خوش آمدید!\n\n"
    "🔧 انجام کلیه خدمات *سخت‌افزاری* و *نرم‌افزاری*\n"
    "🏠 در *منزل* یا محل کار شما، فقط با یک تماس 📞\n"
    "📱 *۰۹۳۶۶۶۴۱۹۶۸*\n\n"
    "لطفاً یکی از گزینه‌های زیر را انتخاب کنید:"
)

# دکمه‌های شیشه‌ای (inline)
inline_keyboard = [
    [InlineKeyboardButton("💼 خدمات", callback_data='services')],
    [InlineKeyboardButton("🎓 آموزش", callback_data='education')],
    [InlineKeyboardButton("📞 تماس", url="tel:+989366641968")],
    [InlineKeyboardButton("🌐 سایت", url="https://yourwebsite.com")],
    [InlineKeyboardButton("📸 اینستاگرام", url="https://instagram.com/yourpage")]
]
inline_markup = InlineKeyboardMarkup(inline_keyboard)

# دکمه‌های ارسال شماره و لوکیشن (reply)
reply_keyboard = [
    [KeyboardButton("📱 ارسال شماره من", request_contact=True)],
    [KeyboardButton("📍 ارسال موقعیت مکانی من", request_location=True)]
]
reply_markup = ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True, one_time_keyboard=True)

# 🔘 /start command
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_markdown(welcome_text, reply_markup=inline_markup)
    await update.message.reply_text("👇 همچنین می‌تونی شماره یا لوکیشن خودتو ارسال کنی:", reply_markup=reply_markup)

# ✳️ دکمه‌های inline
async def handle_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "services":
        await query.message.reply_text(
            "💼 *خدمات ما:*\n"
            "- نصب انواع *ویندوز* و *نرم‌افزارهای تخصصی*\n"
            "- *عیب‌یابی رایگان*\n"
            "- حتی در *ایام تعطیل*! 🛠️",
            parse_mode="Markdown"
        )
    elif query.data == "education":
        await query.message.reply_text(
            "🎓 دوره‌های آموزشی ما:\n"
            "- آموزش کامپیوتر مقدماتی تا پیشرفته\n"
            "- آموزش آفیس، فتوشاپ، ویندوز\n"
            "- پشتیبانی کامل و حضوری"
        )

# 📲 دریافت شماره
async def contact_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    contact = update.message.contact
    user_name = update.effective_user.full_name
    phone = contact.phone_number

    save_to_excel(user_name, phone=phone)
    await update.message.reply_text(f"✅ شماره شما دریافت و ذخیره شد:\n📱 {phone}")

# 📍 دریافت لوکیشن
async def location_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    location = update.message.location
    user_name = update.effective_user.full_name

    save_to_excel(user_name, location=location)
    await update.message.reply_text(
        f"📍 موقعیت شما دریافت و ذخیره شد:\nLatitude: {location.latitude}\nLongitude: {location.longitude}"
    )

# 🚀 اجرا
if __name__ == '__main__':
    TOKEN = os.getenv("BOT_TOKEN")
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(handle_button))
    app.add_handler(MessageHandler(filters.CONTACT, contact_handler))
    app.add_handler(MessageHandler(filters.LOCATION, location_handler))

    print("🤖 ربات در حال اجراست...")
    app.run_polling()